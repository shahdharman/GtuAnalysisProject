from tkinter import *
import sys
import pandas as pd
from tkinter.ttk import *
from tkinter.tix import *
from tkinter.filedialog import *
import shutil,os
import xlsxwriter
import openpyxl as op
from functools import partial
import ntpath
import xlwings as xw
import global_variables as gb 
class GuiApplication:
     #global variables-whose values remain the same for all the objects.
     gb.parent_dir=os.getcwd()#returns currrent directory in which this app is running.


     def __init__(self):#creating folders.
          #global gb.parent_dir,gb.root,gb.l1,gb.l2,gb.e1,gb.e2,input_folder_dirr,output_folder_dirr,gtu_folder_dirr,gb.attendance_folder
          #print("p:",gb.parent_dir)
          gb.root=Tk()#window foundational element on which widgets
          gb.root.title("GTU Analysis Software")
          gb.root.geometry('200x100')
          #self.scrollbar=Scrollbar(gb.root, orient="vertical")
          #self.scrollbar.pack(side=RIGHT,fill='y')
          #self.listbox=Listbox(gb.root,yscrollcommand=self.scrollbar.set)
          #for self.line in range(100):
          #self.listbox.insert(END, "This is line number " + str(self.line))
          #self.listbox.pack(side=LEFT)
          #self.scrollbar.config(command=self.listbox.yview)
          self.path1=os.path.join(gb.parent_dir,"gtufiles")#joins the directory1 with gb.parent_dir as subfolder 
          self.path2=os.path.join(gb.parent_dir,"inputfiles")#joins the directory2 with gb.parent_dir as subfolder
          self.path3=os.path.join(gb.parent_dir,"outputfiles")#joins the directory2 with gb.parent_dir as subfolder
          try:
               os.mkdir(self.path1)
               os.mkdir(self.path2)
               os.mkdir(self.path3)
               gb.gtu_folder_dirr=os.listdir("gtufiles")#list of files or folders in "inputfiles" folder
               gb.output_folder_dirr=os.listdir("outputfiles")#list of files or folders in "outputfiles" folder
               gb.input_folder_dirr=os.listdir("inputfiles")#list of files or folders in "inputfiles" folder
               self.msg = Message(gb.root,text="Folders created succesfully!!")
               self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
               self.msg.pack()
          except:#if folders already exists than pass
               if(os.path.isdir(self.path1)):
                    self.msg=Message(gb.root,text="Folder already exists!!")
                    self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
                    self.msg.pack()
               else:
                    self.msg=Message(gb.root,text="Error!!Folder can't be created")
                    self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
                    self.msg.pack()
      
     def run_app(self):
          print("In run_app function!!")
          gb.l1=Label(gb.root,text="Enter Sem No:")
          gb.l1.pack()#packing label widgetgb.l1 with our window "gb.root"
          gb.e1=Entry(gb.root,bd=5)
          gb.e1.pack()#packing entry widget gb.e1 with our window "gb.root"
          self.btn=Button(gb.root,text='Upload Gtu File',command=self.upload_gtu_file_and_create_subfolders_and_attendance_folder)
          self.btn.pack()
          self.upload_btn=Button(gb.root,text='Upload Attendance File',command= self.upload_attendance_files)
          self.upload_btn.pack()
          #self.l3=Label(gb.root,text="Enter Divisions in order:")
          #self.l3.pack()
          #gb.l2=Label(gb.root,text="Enter Total No. of Divisions:")
          #gb.l2.pack()#packing label widget gb.l2 with our window "gb.root"
          #gb.e2=Entry(gb.root,bd=5)
          #gb.e2.pack()#packing entry widget gb.e2 with our window "gb.root"
          #btn2=Button(gb.root,text='NEXT',command=self.get_divisions)
          #btn2.pack(side=TOP,pady=10)
          self.btn3=Button(gb.root,text='Perform Analysis',command=self.copy_enrollmentNo)
          self.btn3.pack(side=TOP,pady=10)
          self.btn4=Button(gb.root,text='Add data to csv',command=gb.root.quit)
          self.btn4.pack(side=TOP,pady=10)
          gb.root.mainloop()

     #create semname subfolders.
     def create_subfolders(self):
          print("SEm:",gb.e1.get())
          #creating gtufiles subfolder
          self.subfolder="SEM "+ gb.e1.get()
          self.path1=os.path.join("gtufiles",self.subfolder)
          print(self.path1)
          try:
               os.mkdir(self.path1)
               print("gtufiles subfolder created")
          except:
               print("Error! creating gtufiles subfolder!!")
          #creating inputfiles subfolder
          self.subfolder="SEM "+ gb.e1.get()
          self.path=os.path.join("inputfiles",self.subfolder)
          try:
               os.mkdir(self.path)
               print("inputfiles subfolder created")
          except:
              print("Error! creating inputfiles subfolder!") 
          #creating outputfiles subfolder
          self.subfolder="SEM "+ gb.e1.get()
          self.path=os.path.join("outputfiles",self.subfolder)
          try:
               os.mkdir(self.path)
               print("outputfiles subfolder created")
          except:
               print("Error! creating outputfiles subfolder!")
          self.msg=Message(gb.root,text="Sub-Folder created successfully!!")
          self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
          self.msg.pack()
          #os.system('gui.py')


     def upload_gtu_file(self):
          self.dirr="gtufiles"+"/"+"SEM "+ gb.e1.get()
          Tk().withdraw()
          #dst=currdir+"/"+"_SEM"+".xlsx"
          self.filename=askopenfilename(filetypes =[('All Excel Files', '.xlsx .xls')])
          #print(filename)
          #print("Dest:",dirr)
          shutil.copy(self.filename,self.dirr)#copy file from filename(source) to destination(dirr)
          self.filename= ntpath.basename(self.filename)#get last file from absolute path ex file = "/my/little/pony" Ans=pony
          os.replace(self.dirr+"/" + self.filename,self.dirr+"/" + "SEM"+gb.e1.get()+"_gtufile"+".xls")#renaming the uploaded file from current directory
          self.msg = Message(gb.root,text="Gtu file uploaded successfully!!")
          #config is used to access an object's attributes after its initialisation.
          self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
          self.msg.pack()

     #calling both functions     
     def upload_gtu_file_and_create_subfolders_and_attendance_folder(self):
          self.create_subfolders()#creating subfolder with semname
          self.upload_gtu_file()
          #create attendance folder in input files as we only want sem for that
          gb.attendance_folder=os.path.join("inputfiles"+"\\"+"SEM "+ gb.e1.get(),"Attendance Files")#creating Attendance Files folder 
          try:
               os.mkdir(gb.attendance_folder)
          except:
               print("Attendance folder cannot be created.")
          print("path:",gb.attendance_folder)
          
     #to get attendance files for each divisions based on the div_list:
     def upload_attendance_files(self):
         #Tk().withdraw()
          #dst=currdir+"/"+"_SEM"+".xlsx"
          self.filename=askopenfilename(filetypes =[('All Excel Files', '.xlsx .xls')])#returns uploaded of Excel types with .xlsx extension
          shutil.copy(self.filename,gb.attendance_folder)#copy file from filename(source)folder to destination(dirr) folder
          #os.chdir(dirr)#change directory to the Attendance Files Folder.
          self.filename= ntpath.basename(self.filename)#get last file from absolute path ex file = "/my/little/pony" Ans=pony
          os.replace(gb.attendance_folder+"/" + self.filename,gb.attendance_folder+"/" + "SEM"+gb.e1.get()+"_Attendance"+".xlsx")#renaming the uploaded file from current directory
          #print(filename)
          #print("Dest:",dirr)    
          self.msg = Message(gb.root,text="Attendance file uploaded successfully!!")
          #config is used to access an object's attributes after its initialisation.
          self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
          self.msg.pack()
          #os.chdir(gb.parent_dir)
     '''          
     #creating Attendance Folders for div-wise attendance file.
     def get_divisions(self):
          #get divisions and create label widgets and entry widgets based on the no. of divisions.
          #store user input(division) in list.
          global div_list
          self.div_list=[]
          self.divs=gb.e2.get()#returns no. of divisions.
          self.divs=int(self.divs)
          for self.div in range(self.divs):
               gb.l1=Label(gb.root,bd=5,text="Enter div "+str(self.div+1)+":")
               gb.l1.pack()
               self.div_list.append(Entry(gb.root,bd=5))
               self.div_list[-1].pack()
          #create input workbook according to divnames as it's worksheets.    
          
          
     

     #create workbooks in inputfiles and outputfiles for analysis.
     def create_workbook(self):
          print("Parent directory:",gb.parent_dir)
          #Create a workbook and add a worksheet in semname subfolder in inputfiles  folder.
          self.dirr="inputfiles"+"/"+"SEM "+gb.e1.get()+"/"
          os.chdir(self.dirr)#change directory to input files,semname subfolder.
          self.workbook = xlsxwriter.Workbook("inputfile_SEM"+gb.e1.get()+".xlsx")
          for self.div in self.div_list:
               self.worksheet = self.workbook.add_worksheet(self.div.get())#creating worksheets,adds division as it's worksheets, one by one
          self.workbook.close()#must other wise it won't be saved.
          os.chdir(gb.parent_dir)#change to parent directory where this application is present
          # Create a workbook and add a worksheet in semname subfolder in outputfiles  folder.
          self.dirr2="outputfiles"+"/"+"SEM "+gb.e1.get()+"/";
          os.chdir(self.dirr2)#change directory to output files,semname subfolder.
          self.workbook2 = xlsxwriter.Workbook("outputfile_SEM"+gb.e1.get()+".xlsx")#creating workbook
          for self.div in self.div_list:
               self.worksheet2 = self.workbook2.add_worksheet(self.div.get())#adds division as it's worksheets, one by one
          self.msg = Message(gb.root,text="Workbook created successfully!!")
          #config is used to access an object's attributes after its initialisation.
          self.msg.config(bg='lightgreen', font=('times', 10, 'italic'))
          self.msg.pack()
          #Return a list of the worksheet objects in the workbook.
          #btn=Button(gb.root,text="Show Workbook",command=show_Workbook)
          #btn.pack(row=4,column=0,pady=10)
          #The workbook close() method writes all data to the xlsx file and closes it:
          self.workbook2.close()
          os.chdir(gb.parent_dir)
          self.copy_enrollmentNo()

     def show_workbook(self):
          self.book = op.load_workbook("inputfile_SEM"+gb.e1.get()+".xlsx",read_only=True)
          print(self.book.sheetnames)
     '''
     #open both workbooks
     #open worksheet one by one in  both and copy whole "EnrollmentNo"  column in inputfile worksheet
     #copy enrollment Nos from source file(attendance file to our input file.)
     def copy_enrollmentNo(self):
          #source file-->attendance file from which enrollment nos is to be copied(div-wise)
          self.src_workbook=gb.attendance_folder+"/" + "SEM"+gb.e1.get()+"_Attendance"+".xlsx"#source file
          self.dest_workbook="inputfiles"+"/"+"SEM "+gb.e1.get()+"/"+"inputfile_SEM"+gb.e1.get()+".xlsx"#destination file->inputfile of respective sem with divs as their worksheets
          self.src_excel=pd.ExcelFile(self.src_workbook)#returns source excel file.
          self.src_sheets=self.src_excel.sheet_names#list of sheets in source workbook
          self.writer=pd.ExcelWriter(self.dest_workbook,engine="xlsxwriter")#writer object to write dataframe objects into excel file.
          for self.each in self.src_sheets:
               self.src_workbook_df=pd.read_excel(self.src_excel,sheet_name=self.each)#creates dataframe of source file with the sheets
               #main logic starts
               #remove rows with no data
               self.src_workbook_df = self.src_workbook_df.dropna(how = 'all') 
               self.src_workbook_df.columns = [''] * len(self.src_workbook_df.columns)#remove header
               self.src_workbook_df=self.src_workbook_df.iloc[2:]#remove first 2 rows
               self.src_workbook_df.columns=list(self.src_workbook_df.iloc[ 0 , : ].values)#making third row as columns
               self.src_workbook_df.columns=[str(column).replace("\n","") for column in self.src_workbook_df.columns]
               print(self.src_sheets)
               #self.src_workbook_df = self.src_workbook_df['Enrollment\nNo'].replace('\\n',' ')#renaming col from 'Enrollment\nNo' to 'Enrollment No'
               #copying enrollment no from each sheet in "Attendance file" to sheets in 
               #excelwriter->to  write DataFrame objects into excel sheets.
               self.src_workbook_df['EnrollmentNo']=self.src_workbook_df['EnrollmentNo'].astype(str)#converting column to string type
               self.src_workbook_df=self.src_workbook_df[self.src_workbook_df.EnrollmentNo.str.isdigit()]#only numeric values allowed
               self.dest_workbook_df=self.src_workbook_df['EnrollmentNo']#inputfile dataframe
               self.dest_workbook_df.to_excel(self.writer,sheet_name=self.each,index=False)#write to dest workbook
          self.writer.save()#saves our destination workbook
          print("workbook saved successfully")
          print(self.src_sheets)
          self.merge_col_from_gtu_file()
          #main logic ends
          '''
          self.src_wb=op.load_workbook(self.src_workbook,read_only=True)#source file(opening in read only mode as we dont want to change our source file.)
          self.dest_wb =op.load_workbook(self.dest_workbook)#destination file
          
          self.dest_sheets=self.dest_wb.sheetnames#list of sheets in destination workbook
          print(self.src_sheets)
          print("Source Workbook:",self.src_wb)
          print("Destination Workbook:",self.dest_wb)

          for self.each in range(0,len(self.dest_sheets)):#for each worksheet in dest file.
               self.src_sheet =self.src_wb[self.src_sheets[self.each]]#get source sheet from their name
               self.dest_sheet = self.dest_wb[self.dest_sheets[self.each]]#get destination sheet from their name
               print(self.src_sheet)
               print(self.dest_sheet)
               self.counter=0
               for self.i in range(1,self.src_sheet.max_row+1):#these two below nexted loops used to copy all the rows and columns from src to dest.
                    if(self.counter==1):
                         break     #as we only want to fetch first columns                       
                    for self.j in range(1,self.src_sheet.max_column+1):#copy all columns
                         self.dest_sheet.cell(row=self.i, column=self.j).value = self.src_sheet.cell(row=self.i, column=self.j).value
                         self.counter+=1
                         #saving destination worksheet after copying data from source to destination
          self.dest_wb.save(self.dest_workbook)
          '''
     def merge_col_from_gtu_file(self):
          self.src_workbook="gtufiles"+"/"+"SEM "+ gb.e1.get()+"/" + "SEM"+gb.e1.get()+"_gtufile"+".xls"#gtufile workbook path
          self.dest_workbook="inputfiles"+"/"+"SEM "+gb.e1.get()+"/"+"inputfile_SEM"+gb.e1.get()+".xlsx"#destination file->inputfile of respective sem with divs as their worksheets
          self.dest_wb=op.load_workbook(filename = self.dest_workbook)#open inputfile sworkbook to get sheetnames
          self.dest_sheets=self.dest_wb.sheetnames#list of sheets in inputfile workbook
          self.writer=pd.ExcelWriter(self.dest_workbook,engine="openpyxl",mode='a',if_sheet_exists='overlay')#writer object to write dataframe objects into excel file.
          self.writer.book = self.dest_wb
          #self.writer.sheets = dict((ws.title, ws) for ws in self.dest_wb.worksheets)#this is important,other wise the excel writer will create another sheet to write data.
          #self.writer.sheets = {ws.title: ws for ws in self.dest_wb.worksheets}
          for self.each in self.dest_sheets:
               self.src_workbook_df=pd.read_excel(self.src_workbook) #dataframe of gtufile    
               self.dest_workbook_df=pd.read_excel(self.dest_workbook,sheet_name=self.each)#creates dataframe of dest file with the sheets
               #main logic starts
               self.src_workbook_df=self.src_workbook_df.rename(columns={'UNIT_NO':'EnrollmentNo'})#renaming unitno to 'Enrollment No' for merging
               print("Source:",self.src_workbook_df.columns)
               print("Destination",self.dest_workbook_df.columns)
               #merging all the columns in inputfile(div-wise) based on enrollment no 
               self.dest_workbook_df=self.dest_workbook_df.merge(self.src_workbook_df,on='EnrollmentNo')
               print("dataframe:",self.dest_workbook_df)
               self.dest_workbook_df.to_excel(self.writer,sheet_name=self.each,index=False)#write to dest workbook
               #self.dest_workbook_df.to_csv(self.dest_workbook,mode='a', index = False, header=None)
          self.writer.save()#saves our destination workbook
          print("workbook saved successfully")

     def get_subject_dict(self):
          #Get a dictionary of subjectname,subjectoverall and subject_individual marks(ie SUB(N)NA,SUB(N)GR,SUB(N)B columns)
          self.gtu_workbook="gtufiles"+"/"+"SEM "+ gb.e1.get()+"/" + "SEM"+gb.e1.get()+"_gtufile"+".xls"#gtufile workbook path
          self.df=pd.read_excel(self.gtu_workbook)
          self.column_dict=self.df.to_dict('dict')
          #location of column "SUB1NA"
          key_loc1= pd.Index(self.column_dict).get_loc('SUB1NA')

          #location of column "SUB1CR"(SUB1CR because it is the next column after SUB(N)NA,SUB(N)GR,SUB(N)B columns.)
          key_loc2= pd.Index(self.column_dict).get_loc('SUB1GRI')
          
          #fill 0 in place of nan values.  
          self.df.fillna(0,inplace=True)

          # Remove all columns between column index 0 to key_loc(excluding key_loc)
          self.df.drop(self.df.iloc[:, 0:key_loc1], inplace = True, axis = 1)
          
          # Remove all columns between from key_loc2 onwards,hence now only subject columns are left.
          self.df.drop(self.df.iloc[:, key_loc2:], inplace = True, axis = 1)
          self.column_dict=self.df.to_dict('dict')
          print(self.df.columns)

          #location of column "SUB1GR"(which gives overall subject result )
          key_loc3= pd.Index(self.column_dict).get_loc('SUB1GR')

          #location of column "SUB1B"(which gives theory,practical results.)
          key_loc4= pd.Index(self.column_dict).get_loc('SUB1B')
          
          #adding subjects to the dict.
          count=1
          for key,value in self.column_dict.items():
               if key[-3:len(key)]==str(count)+"NA" and value[0]!=0:
                    gb.sub_dict[key]=[value[0]]
                    count+=1
               else:
                    break
               
          count=1    
          for sub in gb.sub_dict.keys():
               for key,value in self.column_dict.items():
                    if key[-3:len(key)]==str(count)+"GR":
                         gb.sub_dict[sub].append(key)
                         count+=1
                         break
          count=1
          for sub in gb.sub_dict.keys():
               for key,value in self.column_dict.items():
                    if key[-2:len(key)]==str(count)+"B":
                         gb.sub_dict[sub].append(key)
                         count+=1
                         break
                             #List of subjects.
          print("Subject dictionary:",gb.sub_dict)
          #get list of subjects from the gtu file
          #by iterating columns from the col "SUB1NA" till "SUB(N)NA"(N=no of subs) till the col has value
          #gb.sub_dict=
